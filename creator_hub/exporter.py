from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from .db import connect, json_load
from .util import now_utc


def _video_rows(conn):
    rows=conn.execute("""SELECT v.*,c.channel_title,c.country_api,c.subscriber_count,
      s.suggested_role,s.brands_json AS suggested_brands_json,s.confidence,s.evidence_json,
      l.human_role,l.brands_json AS human_brands_json,l.labeled_by,l.note AS label_note,l.labeled_at
      FROM videos v JOIN creators c ON c.channel_id=v.channel_id
      LEFT JOIN label_suggestions s ON s.video_id=v.video_id
      LEFT JOIN video_labels l ON l.video_id=v.video_id
      ORDER BY v.channel_id,v.published_at DESC""").fetchall()
    out=[]
    for r in rows:
        d=dict(r)
        d["tags"]=json_load(d.pop("tags_json"),[])
        d["suggested_brands"]=json_load(d.pop("suggested_brands_json"),[])
        d["suggestion_evidence"]=json_load(d.pop("evidence_json"),[])
        d["human_brands"]=json_load(d.pop("human_brands_json"),[])
        out.append(d)
    return out



def xlsx_bytes(sheet_name: str, columns: list[tuple[str,str]], rows, *, metadata: list[tuple[str,Any]] | None = None, extra_sheets: list[tuple[str,list[tuple[str,str]],Any]] | None = None) -> bytes:
    """Build a memory-efficient XLSX from an iterable of dictionaries."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except Exception as e:
        raise RuntimeError("XLSX 导出需要 openpyxl。执行 python -m pip install -r requirements.txt") from e
    from io import BytesIO
    wb=Workbook(write_only=True)
    ws=wb.create_sheet((sheet_name or 'Data')[:31])
    ws.append([label for _,label in columns])
    for row in rows:
        vals=[]
        for key,_ in columns:
            v=row.get(key) if isinstance(row,dict) else None
            if isinstance(v,(list,dict)): v=json.dumps(v,ensure_ascii=False)
            vals.append(v)
        ws.append(vals)
    if metadata:
        meta=wb.create_sheet('Export Info')
        meta.append(['Field','Value'])
        for k,v in metadata:
            meta.append([k,json.dumps(v,ensure_ascii=False) if isinstance(v,(list,dict)) else v])
    for extra_name, extra_columns, extra_rows in (extra_sheets or []):
        ews=wb.create_sheet((extra_name or 'Extra')[:31])
        ews.append([label for _,label in extra_columns])
        for row in extra_rows:
            vals=[]
            for key,_ in extra_columns:
                v=row.get(key) if isinstance(row,dict) else None
                if isinstance(v,(list,dict)): v=json.dumps(v,ensure_ascii=False)
                vals.append(v)
            ews.append(vals)
    bio=BytesIO(); wb.save(bio); return bio.getvalue()


def safe_export_filename(value: str, fallback: str='export') -> str:
    import re
    x=re.sub(r'[^A-Za-z0-9._-]+','_',str(value or '')).strip('._')
    return (x or fallback)[:120]

def export_all(db_path: str|Path, output_dir: str|Path, fmt: str="csv") -> dict[str,Any]:
    out=Path(output_dir); out.mkdir(parents=True,exist_ok=True)
    stamp=now_utc().replace(":","").replace("-","")[:15]
    with connect(db_path) as conn:
        creators=[dict(r) for r in conn.execute("SELECT * FROM creators ORDER BY channel_title").fetchall()]
        videos=_video_rows(conn)
        snapshots=[dict(r) for r in conn.execute("SELECT * FROM video_snapshots ORDER BY video_id,captured_at").fetchall()]
        discovery_runs=[dict(r) for r in conn.execute("SELECT * FROM discovery_runs ORDER BY started_at").fetchall()]
        discovery_creators=[dict(r) for r in conn.execute("SELECT * FROM discovery_creator_results ORDER BY run_id,id").fetchall()]
        discoveries=[dict(r) for r in conn.execute("SELECT * FROM discovery_hits ORDER BY id").fetchall()]
        labels=[dict(r) for r in conn.execute("SELECT * FROM video_labels ORDER BY labeled_at").fetchall()]
        audits=[dict(r) for r in conn.execute("SELECT * FROM video_label_audit ORDER BY id").fetchall()]
        workflows=[dict(r) for r in conn.execute("SELECT * FROM creator_workflow ORDER BY updated_at").fetchall()]
        workflow_audit=[dict(r) for r in conn.execute("SELECT * FROM creator_workflow_audit ORDER BY id").fetchall()]
        discovery_summary=[dict(r) for r in conn.execute("SELECT * FROM creator_discovery_summary ORDER BY last_seen_at").fetchall()]
        sync_attempts=[dict(r) for r in conn.execute("SELECT * FROM creator_sync_attempts ORDER BY id").fetchall()]
        maintenance=[dict(r) for r in conn.execute("SELECT * FROM maintenance_runs ORDER BY id").fetchall()]
        app_settings=[dict(r) for r in conn.execute("SELECT * FROM app_settings ORDER BY key").fetchall()]
    data={"creators":creators,"videos":videos,"video_snapshots":snapshots,"discovery_runs":discovery_runs,"discovery_creator_results":discovery_creators,"discovery_hits":discoveries,"human_labels":labels,"label_audit":audits,"creator_workflow":workflows,"creator_workflow_audit":workflow_audit,"creator_discovery_summary":discovery_summary,"creator_sync_attempts":sync_attempts,"maintenance_runs":maintenance,"app_settings":app_settings}
    if fmt=="json":
        p=out/f"creator_data_hub_{stamp}.json"; p.write_text(json.dumps(data,ensure_ascii=False,indent=2),encoding="utf-8")
        return {"format":"json","files":[str(p.resolve())]}
    if fmt=="xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except Exception as e:
            raise RuntimeError("XLSX 导出需要 openpyxl。执行 python -m pip install -r requirements.txt") from e
        p=out/f"creator_data_hub_{stamp}.xlsx"; wb=Workbook(); wb.remove(wb.active)
        for name, rows in [("Creators",creators),("Videos",videos),("Video Snapshots",snapshots),("Discovery Runs",discovery_runs),("Discovery Creators",discovery_creators),("Discovery Hits",discoveries),("Human Labels",labels),("Label Audit",audits),("Creator Workflow",workflows),("Workflow Audit",workflow_audit),("Discovery Summary",discovery_summary),("Sync Attempts",sync_attempts),("Maintenance",maintenance),("App Settings",app_settings)]:
            ws=wb.create_sheet(name[:31])
            if not rows:
                ws.append(["No data"]); continue
            headers=list(rows[0].keys()); ws.append(headers)
            for c in ws[1]: c.font=Font(bold=True)
            for row in rows:
                ws.append([json.dumps(row.get(h),ensure_ascii=False) if isinstance(row.get(h),(list,dict)) else row.get(h) for h in headers])
            ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
            for col in ws.columns:
                maxlen=min(50,max(len(str(c.value or "")) for c in col)); ws.column_dimensions[col[0].column_letter].width=max(10,maxlen+2)
        ws=wb.create_sheet("Data Dictionary")
        ws.append(["层级","字段/表","说明"])
        for row in [
            ("事实","creators","YouTube频道公开事实与当前快照"),("事实","videos","公开视频元数据与当前Views/Likes/Comments"),("事实","video_snapshots","每次真实刷新时的指标快照"),("发现", "discovery_runs", "一次搜索批次的关键词、Query Expansion、时间/地区条件和完成状态"),("发现", "discovery_creator_results", "一次搜索批次 × 一个博主的去重结果、最佳命中、发现评分和Query Coverage"),("发现", "discovery_hits", "一次搜索批次内每个实际Query命中的视频证据"),("机器标签","label_suggestions","根据公开metadata给出的建议，非原始事实"),("人工标签","video_labels","运营确认结果"),("审计","video_label_audit","人工标签变更历史")
        ]: ws.append(row)
        wb.save(p); return {"format":"xlsx","files":[str(p.resolve())]}
    files=[]
    for name,rows in data.items():
        p=out/f"{name}_{stamp}.csv"; files.append(str(p.resolve()))
        headers=list(rows[0].keys()) if rows else ["no_data"]
        with p.open("w",encoding="utf-8-sig",newline="") as f:
            w=csv.DictWriter(f,fieldnames=headers); w.writeheader()
            for row in rows:
                w.writerow({k: json.dumps(v,ensure_ascii=False) if isinstance(v,(list,dict)) else v for k,v in row.items()})
    return {"format":"csv","files":files}
