from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from .db import connect, json_dump
from .service import CreatorHub
from .util import now_utc, parse_duration_seconds


def _read_jsonl(path: Path):
    with path.open("r", encoding="utf-8", errors="replace") as f:
        for line in f:
            line=line.strip()
            if not line: continue
            try: yield json.loads(line)
            except Exception: continue


def import_v2(hub: CreatorHub, root: str | Path, *, monitoring: bool = True) -> dict[str, Any]:
    root = Path(root)
    creator_meta_files = list(root.rglob("channel_metadata.json"))
    creators = videos = snapshots = 0
    notes: list[str] = []
    for meta_file in creator_meta_files:
        folder = meta_file.parent
        try:
            meta = json.loads(meta_file.read_text(encoding="utf-8"))
        except Exception as e:
            notes.append(f"skip {meta_file}: {e}")
            continue
        cid = meta.get("channel_id") or meta.get("id")
        if not cid:
            notes.append(f"skip {meta_file}: missing channel_id")
            continue
        row = {
            "channel_id": cid,
            "channel_title": meta.get("title") or meta.get("channel_title") or "",
            "handle": meta.get("custom_url") or meta.get("handle") or "",
            "channel_url": meta.get("canonical_url") or f"https://www.youtube.com/channel/{cid}",
            "description": meta.get("description") or "",
            "country_api": meta.get("country") or meta.get("country_api") or "",
            "published_at": meta.get("published_at") or "",
            "subscriber_count": meta.get("subscriber_count"),
            "channel_view_count": meta.get("channel_view_count"),
            "channel_video_count": meta.get("channel_video_count"),
            "hidden_subscriber_count": 1 if meta.get("hidden_subscriber_count") else 0,
            "uploads_playlist_id": meta.get("uploads_playlist") or meta.get("uploads_playlist_id") or "",
            "thumbnail_url": meta.get("thumbnail_url") or "",
        }
        hub.upsert_creator(row, monitoring=monitoring, source="v2_import", snapshot=True)
        creators += 1
        vfile = folder / "videos_classified.jsonl"
        if vfile.exists():
            for item in _read_jsonl(vfile):
                vid = item.get("video_id")
                if not vid: continue
                collected = item.get("collected_at_utc") or item.get("last_metric_at") or now_utc()
                vrow = {
                    "video_id": vid,
                    "channel_id": item.get("channel_id") or cid,
                    "title": item.get("title") or "",
                    "description": item.get("description") or "",
                    "tags": item.get("tags") or [],
                    "published_at": item.get("published_at") or "",
                    "duration_iso8601": item.get("duration_iso8601") or item.get("duration_iso") or "",
                    "duration_seconds": item.get("duration_seconds") if item.get("duration_seconds") is not None else parse_duration_seconds(item.get("duration_iso8601")),
                    "live_broadcast_content": item.get("live_broadcast_content") or "none",
                    "category_id": item.get("category_id") or "",
                    "default_language": item.get("default_language") or "",
                    "privacy_status": item.get("privacy_status") or "",
                    "thumbnail_url": item.get("thumbnail_url") or "",
                    "current_views": item.get("views"),
                    "current_likes": item.get("likes"),
                    "current_comments": item.get("comments"),
                    "last_metric_at": collected,
                    "discovered_at": collected,
                }
                # Legacy classification is imported only as a machine suggestion, never as a human label.
                legacy_class = item.get("classification")
                mapping = {"unrelated":"daily", "ugphone":"ugphone", "competitor":"competitor", "multi_brand_cloud_phone":"multi_brand", "other_cloud_phone":"other_cloud_phone"}
                suggestion = None
                if legacy_class:
                    suggestion = {
                        "video_id": vid,
                        "suggested_role": mapping.get(legacy_class, "pending"),
                        "brands": item.get("matched_brands") or item.get("strong_matched_brands") or [],
                        "confidence": {"confirmed":"high", "probable":"medium", "review":"review"}.get(item.get("classification_confidence"), "review"),
                        "evidence": ["legacy_v2_suggestion"] + list(item.get("evidence") or []),
                        "generated_at": collected,
                        "rule_version": "legacy-v2",
                    }
                hub.upsert_video(vrow, snapshot=True, suggestion=suggestion)
                videos += 1
        sfile = folder / "video_snapshots.jsonl"
        if sfile.exists():
            with connect(hub.db_path) as conn:
                for item in _read_jsonl(sfile):
                    vid = item.get("video_id")
                    at = item.get("collected_at_utc") or item.get("captured_at")
                    if not vid or not at: continue
                    exists = conn.execute("SELECT video_id FROM videos WHERE video_id=?", (vid,)).fetchone()
                    if not exists: continue
                    cur = conn.execute("INSERT OR IGNORE INTO video_snapshots(video_id,captured_at,views,likes,comments) VALUES(?,?,?,?,?)",
                                       (vid,at,item.get("views"),item.get("likes"),item.get("comments")))
                    if cur.rowcount:
                        snapshots += 1
                conn.commit()
    with connect(hub.db_path) as conn:
        conn.execute("INSERT INTO imports(source_type,source_path,imported_at,creators,videos,snapshots,message) VALUES(?,?,?,?,?,?,?)",
                     ("youtube-kol-gmv-intelligence-v2", str(root.resolve()), now_utc(), creators, videos, snapshots, "\n".join(notes)[:10000]))
        conn.commit()
    business={"metric_values_upserted":0,"creators_matched":0,"unmatched_rows":0}
    try:
        business=import_business_metrics(hub,root,source_type="youtube-kol-gmv-intelligence-v2")
    except Exception as e:
        notes.append(f"business metrics scan skipped: {type(e).__name__}: {e}")
    return {"creators":creators,"videos":videos,"snapshots_added":snapshots,"business_metrics":business,"notes":notes}

# ---------- creator business metrics (GMV / acquisition / future backend feeds) ----------
import csv
import re
import uuid

_ID_ALIASES={
    "channel_id","channel id","youtube channel id","youtube_channel_id","频道id","频道 id","博主id","博主 id"
}
_HANDLE_ALIASES={"handle","youtube handle","频道handle","博主handle","账号","博主账号","youtube账号","youtube 账号","kol账号","kol handle"}
_URL_ALIASES={"channel_url","channel url","youtube url","youtube_url","频道链接","youtube链接","博主链接","频道url","youtube主页","youtube 主页","博主主页","频道主页","主页链接"}
_TITLE_ALIASES={"channel_title","channel title","channel name","creator","creator name","博主","博主名称","博主名","频道名称","频道名","频道","kol","kol名称","kol name","creator名称"}
_METRIC_ALIASES={
    "gmv":{"gmv","累计gmv","总gmv","gmv金额","gmv usd","gmv($)","gmv（$）","gmv(usd)","gmv（usd）","销售额","成交额","商品交易总额","gross merchandise value"},
    "new_users":{"new_users","new users","new user","new user count","拉新","拉新数","拉新人数","拉新用户","拉新用户数","新增用户","新增用户数","新增人数","新用户","新用户数","acquisition","acquisitions","acquired users"},
    "orders":{"orders","order count","订单","订单数"},
    "revenue":{"revenue","营收","收入"},
    "commission":{"commission","佣金","佣金金额"},
    "cost":{"cost","合作费用","费用","投放成本"},
}
_PERIOD_START={"period_start","start_date","from_date","开始日期","周期开始","统计开始"}
_PERIOD_END={"period_end","end_date","to_date","结束日期","周期结束","统计结束"}
_CURRENCY={"currency","币种","货币"}
_CAMPAIGN={"campaign","campaign_name","活动","活动名称","项目","项目名称"}
_REGION={"region","country","国家","地区","市场"}
_NOTE={"note","备注","说明"}


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().casefold()).replace("-", "_")


def _first_col(headers: list[str], aliases: set[str]) -> int | None:
    norm_alias={_norm_header(x) for x in aliases}
    for i,h in enumerate(headers):
        if _norm_header(h) in norm_alias:
            return i
    return None


def _metric_cols(headers: list[str]) -> dict[int,str]:
    out={}
    normalized=[_norm_header(h) for h in headers]
    for key,aliases in _METRIC_ALIASES.items():
        aset={_norm_header(x) for x in aliases}
        for i,h in enumerate(normalized):
            if h in aset:
                out[i]=key
    return out


def _float_value(value: Any) -> float | None:
    if value is None or value=="": return None
    if isinstance(value,(int,float)): return float(value)
    s=str(value).strip().replace(",","").replace("¥","").replace("$","").replace("￥","")
    if not s or s in {"—","-","n/a","nan","none"}: return None
    try: return float(s)
    except Exception: return None


def _text(value: Any) -> str:
    if value is None: return ""
    # openpyxl may surface date/datetime objects; ISO-ish text is best for lineage.
    if hasattr(value,"isoformat"):
        try: return value.isoformat()[:10]
        except Exception: pass
    return str(value).strip()


def _iter_business_sheets(path: Path):
    suffix=path.suffix.lower()
    if suffix==".csv":
        last=None
        for enc in ("utf-8-sig","utf-8","gb18030"):
            try:
                with path.open("r",encoding=enc,newline="") as f:
                    rows=list(csv.reader(f))
                if rows: yield path.name,rows
                return
            except Exception as e: last=e
        if last: raise last
    elif suffix in {".xlsx",".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as e:
            raise RuntimeError("openpyxl is required to import XLSX business metrics") from e
        wb=load_workbook(path,read_only=True,data_only=True)
        try:
            for ws in wb.worksheets:
                rows=[[c for c in row] for row in ws.iter_rows(values_only=True)]
                if rows: yield ws.title,rows
        finally:
            wb.close()


def _creator_match_indexes(conn):
    rows=conn.execute("SELECT channel_id,channel_title,handle,channel_url FROM creators").fetchall()
    by_id={str(r["channel_id"]):str(r["channel_id"]) for r in rows}
    by_handle={}
    by_title={}
    for r in rows:
        cid=str(r["channel_id"])
        h=str(r["handle"] or "").strip().casefold().lstrip("@")
        if h: by_handle.setdefault(h,[]).append(cid)
        t=" ".join(str(r["channel_title"] or "").split()).casefold()
        if t: by_title.setdefault(t,[]).append(cid)
    return by_id,by_handle,by_title


def _resolve_business_creator(row: list[Any], headers: list[str], indexes, cols) -> tuple[str|None,str]:
    by_id,by_handle,by_title=indexes
    def at(i): return _text(row[i]) if i is not None and i<len(row) else ""
    raw_id=at(cols["id"])
    if raw_id:
        m=re.search(r"(UC[\w-]{20,})",raw_id)
        cid=(m.group(1) if m else raw_id.strip())
        if cid in by_id: return cid,"channel_id"
    url=at(cols["url"])
    if url:
        m=re.search(r"youtube\.com/channel/(UC[\w-]{20,})",url,re.I)
        if m and m.group(1) in by_id: return m.group(1),"channel_url"
        hm=re.search(r"youtube\.com/@([^/?#]+)",url,re.I)
        if hm:
            vals=by_handle.get(hm.group(1).casefold(),[])
            if len(vals)==1:return vals[0],"channel_url_handle"
    handle=at(cols["handle"]).lstrip("@").casefold()
    if handle:
        vals=by_handle.get(handle,[])
        if len(vals)==1:return vals[0],"handle"
    title=" ".join(at(cols["title"]).split()).casefold()
    if title:
        vals=by_title.get(title,[])
        if len(vals)==1:return vals[0],"channel_title_exact"
    return None,"unmatched"


def import_business_metrics(hub: CreatorHub, source: str|Path, *, source_type: str="manual_import", progress=None) -> dict[str,Any]:
    """Import creator-grain commercial facts without coupling them to the creators table.

    CSV/XLSX files are accepted. A sheet is considered relevant only when it contains at
    least one creator identity column and at least one known business metric column. Rows
    are matched deterministically against already-known Creator records; unmatched rows
    are reported and never guessed into the database.
    """
    source=Path(source)
    source_is_dir=source.is_dir()
    files=[]
    if source_is_dir:
        files=sorted([p for p in source.rglob("*") if p.is_file() and p.suffix.lower() in {".csv",".xlsx",".xlsm"}])
    elif source.is_file(): files=[source]
    else: raise FileNotFoundError(str(source))
    batch="biz-"+uuid.uuid4().hex[:12]
    imported=updated=skipped=matched_rows=0; touched=set(); notes=[]; relevant_files=0
    total=max(1,len(files))
    with connect(hub.db_path) as conn:
        indexes=_creator_match_indexes(conn)
        for fi,path in enumerate(files,1):
            if progress: progress(stage="商业数据导入",message=f"正在扫描 {path.name}",current=fi-1,total=total,percent=round((fi-1)*100/total,1))
            try:
                sheets=list(_iter_business_sheets(path))
            except Exception as e:
                notes.append(f"skip {path}: {type(e).__name__}: {e}"); continue
            file_relevant=False
            for sheet,rows in sheets:
                # Search the first 15 rows for a usable header. Old workbooks often have a title row.
                header_i=None; meta=None
                for hi,candidate in enumerate(rows[:15]):
                    headers=[_text(x) for x in candidate]
                    mcols=_metric_cols(headers)
                    cols={"id":_first_col(headers,_ID_ALIASES),"handle":_first_col(headers,_HANDLE_ALIASES),"url":_first_col(headers,_URL_ALIASES),"title":_first_col(headers,_TITLE_ALIASES)}
                    if mcols and any(v is not None for v in cols.values()):
                        header_i=hi;meta=(headers,mcols,cols);break
                if meta is None: continue
                file_relevant=True
                headers,mcols,cols=meta
                extra={
                    "start":_first_col(headers,_PERIOD_START),"end":_first_col(headers,_PERIOD_END),"currency":_first_col(headers,_CURRENCY),
                    "campaign":_first_col(headers,_CAMPAIGN),"region":_first_col(headers,_REGION),"note":_first_col(headers,_NOTE),
                }
                for ri,row in enumerate(rows[header_i+1:],header_i+2):
                    cid,matched_by=_resolve_business_creator(list(row),headers,indexes,cols)
                    values=[(idx,key,_float_value(row[idx] if idx<len(row) else None)) for idx,key in mcols.items()]
                    values=[x for x in values if x[2] is not None]
                    if not values: continue
                    if not cid:
                        skipped+=1
                        if len(notes)<100: notes.append(f"unmatched {path.name}/{sheet} row {ri}")
                        continue
                    matched_rows+=1;touched.add(cid)
                    def at(i): return _text(row[i]) if i is not None and i<len(row) else ""
                    common=(at(extra["currency"]),at(extra["start"]),at(extra["end"]),at(extra["campaign"]),at(extra["region"]),at(extra["note"]))
                    for idx,key,val in values:
                        currency=common[0]
                        if not currency and key in {"gmv","revenue","commission","cost"}:
                            hn=_norm_header(headers[idx])
                            if "usd" in hn or "$" in headers[idx]: currency="USD"
                            elif "cny" in hn or "rmb" in hn or "¥" in headers[idx] or "￥" in headers[idx]: currency="CNY"
                        try: stable_file=str(path.relative_to(source)) if source_is_dir else path.name
                        except Exception: stable_file=path.name
                        source_ref=f"{stable_file}::{sheet}::{ri}::{headers[idx]}"
                        raw={headers[i]:_text(row[i]) for i in range(min(len(headers),len(row))) if headers[i]}
                        before=conn.total_changes
                        conn.execute("""INSERT INTO creator_business_metrics(channel_id,metric_key,metric_value,currency,period_start,period_end,campaign,region,source_type,source_ref,import_batch,captured_at,note,raw_json)
                                      VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                                      ON CONFLICT(channel_id,metric_key,period_start,period_end,campaign,region,source_type,source_ref)
                                      DO UPDATE SET metric_value=excluded.metric_value,currency=excluded.currency,import_batch=excluded.import_batch,captured_at=excluded.captured_at,note=excluded.note,raw_json=excluded.raw_json""",
                                     (cid,key,float(val),currency,common[1],common[2],common[3],common[4],source_type,source_ref,batch,now_utc(),common[5],json_dump({"matched_by":matched_by,"row":raw})))
                        # SQLite total_changes increments for both insert and update; expose as imported/upserted.
                        imported+=1
            if file_relevant: relevant_files+=1
            if progress: progress(stage="商业数据导入",message=f"已扫描 {fi}/{total} 个文件 · 已写入 {imported} 个指标值",current=fi,total=total,percent=round(fi*100/total,1))
        conn.execute("INSERT INTO imports(source_type,source_path,imported_at,creators,videos,snapshots,message) VALUES(?,?,?,?,?,?,?)",
                     ("creator-business-metrics",str(source.resolve()),now_utc(),len(touched),0,0,json_dump({"batch":batch,"metric_values":imported,"matched_rows":matched_rows,"unmatched_rows":skipped,"relevant_files":relevant_files,"notes":notes[:100]})))
        conn.commit()
    return {"ok":True,"import_batch":batch,"source":str(source.resolve()),"files_scanned":len(files),"relevant_files":relevant_files,"creators_matched":len(touched),"rows_matched":matched_rows,"metric_values_upserted":imported,"unmatched_rows":skipped,"notes":notes}
